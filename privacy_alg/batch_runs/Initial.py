import numpy as np
import matplotlib.pyplot as plt
import networkx as nx
from scipy.linalg import eigvals, eig
from scipy.optimize import minimize
import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ------------------------------------------------------------
# Row-stochastic normalization
def row_normalize(M: np.ndarray) -> np.ndarray:
    s = M.sum(axis=1, keepdims=True)
    s[s == 0] = 1e-12
    return M / s

# Build augmented P with fixed gadget weights
def build_Ap(N: int, A: np.ndarray) -> np.ndarray:
    size = 4 * N
    Ap = np.zeros((size, size))
    Ap[:N, :N] = A
    for i in range(N):
        inds = {
            0: i,        
            1: N + 3*i,   
            2: N + 3*i+1, 
            3: N + 3*i+2  
        }

        # ——————————————————————
        # a)
        
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[3]] = 2
        Ap[inds[3], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        
        # ——————————————————————
        # b)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # c)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[1], inds[3]] = 1
        Ap[inds[2], inds[3]] = 1
        Ap[inds[3], inds[0]] = 2
        '''
        # ——————————————————————
        # d)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        '''
        # ——————————————————————
        # e)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[0]] = 1
        '''
        # ——————————————————————
        # f)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[1], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[3]] = 2
        Ap[inds[3], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # g)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # h)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[1], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # i)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[2]] = 1
        Ap[inds[2], inds[1]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # j)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[2], inds[1]] = 1
        Ap[inds[2], inds[3]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[0]] = 1
        '''
        # ——————————————————————
        # k)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[1], inds[2]] = 2
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[3], inds[0]] = 1
        '''
        # ——————————————————————
        # l)
        '''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[0]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[2], inds[1]] = 1
        Ap[inds[2], inds[3]] = 1
        '''
        # ——————————————————————
        # m)
        ''''
        Ap[inds[0], inds[1]] = 1
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[0]] = 1
        Ap[inds[0], inds[3]] = 1
        Ap[inds[1], inds[2]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
    return row_normalize(Ap)

def build_A_from_p(p: np.ndarray, mask: np.ndarray) -> np.ndarray:
    N = mask.shape[0]
    A = np.zeros((N, N))
    idx = 0
    for i in range(N):
        for j in range(N):
            if mask[i, j]:
                A[i, j] = p[idx]
                idx += 1
    return row_normalize(A)

def simulate_and_metrics(A_mat: np.ndarray,
                         x0: np.ndarray,
                         tol: float = 1e-4,
                         max_steps: int = 500) -> dict:
    N = x0.size
    P_mat = build_Ap(N, A_mat)

    # second‐largest eigenvalue
    eigs = np.abs(eigvals(P_mat))
    eigs.sort()
    lambda2 = eigs[-2]

    # left stationary eigenvector → v0
    wv, V = eig(P_mat.T)
    idx = np.argmin(np.abs(wv - 1))
    v0 = np.real(V[:, idx])
    v0 /= v0.sum()

    # build augmented x0
    target = x0.mean()
    x_p0 = np.zeros(4*N)
    for j in range(N):
        x_p0[N+3*j:N+3*j+3] = 4 * x0[j] / 3
    x_p0 *= target / (v0 @ x_p0)

    # simulate both systems
    Xo = np.zeros((N,   max_steps+1))
    Xa = np.zeros((4*N, max_steps+1))
    Xo[:, 0] = x0
    Xa[:, 0] = x_p0

    orig_steps = None
    aug_steps  = None

    for k in range(max_steps):
        Xo[:, k+1] = A_mat @ Xo[:, k]
        Xa[:, k+1] = P_mat @ Xa[:, k]
        if orig_steps is None and np.max(np.abs(Xo[:, k+1] - target)) < tol:
            orig_steps = k+1
        if aug_steps  is None and np.max(np.abs(Xa[:, k+1] - target)) < tol:
            aug_steps = k+1

    return {
        'lambda2':    lambda2,
        'orig_steps': orig_steps,
        'aug_steps':  aug_steps
    }

# ------------------------------------------------------------
# 2) Closure‐factory for the optimizer

def make_obj(N: int, mask: np.ndarray):
    def objective(p):
        A = build_A_from_p(p, mask)
        vals = np.abs(eigvals(build_Ap(N, A)))
        vals.sort()
        return vals[-2]
    return objective

# ------------------------------------------------------------
# 3) Single‐experiment runner

def one_experiment(N: int,
                   seed: int,
                   p_edge: float,
                   x0: np.ndarray,
                   tol: float = 1e-4,
                   max_steps: int = 500) -> dict:

    rng = np.random.RandomState(seed)

    while True:
        D = nx.gnp_random_graph(N,
                                p_edge,
                                seed=rng,
                                directed=True)
        if nx.is_strongly_connected(D):
            G = D
            break

    # 3) Assign random weights from the same RNG
    for u, v in G.edges():
        G[u][v]['weight'] = rng.rand()

    # 4) Build the row‐normalized adjacency A0
    raw_A = row_normalize(nx.to_numpy_array(G, nodelist=range(N), weight='weight'))
    mask  = raw_A > 0
    p0    = np.array([raw_A[i, j]
                      for i in range(N) for j in range(N)
                      if mask[i, j]])
    A0    = build_A_from_p(p0, mask)

    # 5) Simulate the original
    init_m = simulate_and_metrics(A0, x0, tol, max_steps)

    # 6) Optimize only the free entries
    obj = make_obj(N, mask)
    res = minimize(obj,
                   p0,
                   method='SLSQP',
                   bounds=[(0.1, None)] * p0.size,
                   options={'ftol': 1e-9, 'maxiter': 200})

    # 7) Build & simulate the optimized A
    A_opt = build_A_from_p(res.x, mask)
    opt_m = simulate_and_metrics(A_opt, x0, tol, max_steps)

    return {
        'run':            seed,       
        'A0':             A0,          
        'lambda2_init':   init_m['lambda2'],
        'lambda2_opt':    opt_m['lambda2'],
        'steps_init':     init_m['orig_steps'],
        'steps_opt':      opt_m['orig_steps'],
        'aug_steps_init': init_m['aug_steps'],
        'aug_steps_opt':  opt_m['aug_steps'],
        'success':        res.success
    }

if __name__ == "__main__":
    N        = 50
    p_edge   = 0.3
    num_runs = 70

    #11x11
    #x0_common = np.array([0.1,0.3,0.6,0.43,0.85,0.9,0.45,0.11,0.06,0.51,0.13])

    #20x20
    #x0_common = np.array([0.1, 0.3, 0.6, 0.43, 0.85, 0.9, 0.45, 0.11, 0.06, 0.51, 0.13, 0.27, 0.72, 0.14, 0.99, 0.34, 0.58, 0.21, 0.47, 0.80])
    
    #50x50
    rng = np.random.RandomState(1234)
    x0_common = rng.rand(50)
    results = []
    all_A0 = []



    # 1) Run experiments
    for run in range(num_runs):
        seed = 4 + run
        m = one_experiment(N, seed, p_edge, x0_common,
                           tol=1e-4, max_steps=1000)
        m['run'] = run+1
        # compute % speedup
        if m['aug_steps_init'] and m['aug_steps_opt']:
            m['pct_speedup'] = 100*(m['aug_steps_init']-m['aug_steps_opt'])/m['aug_steps_init']
        else:
            m['pct_speedup'] = np.nan
        results.append(m)
        all_A0.append(m['A0'])         # stash each A0
        print(f"Run {run+1}/{num_runs}: P_init={m['aug_steps_init']}, "
              f"P_opt={m['aug_steps_opt']}, speedup={m['pct_speedup']:.1f}%")

    # 2) Build summary DataFrame
    df = pd.DataFrame(results)[[
        'run','lambda2_init','lambda2_opt',
        'aug_steps_init','aug_steps_opt','pct_speedup'
    ]]

    # 3) Ensure output folder
    out_dir = Path("ExcelDataRuns")
    out_dir.mkdir(exist_ok=True)

    # 4) Save summary + boxplot
    excel_path = out_dir/"InitialA_N=50_P=0.3.xlsx"
    df.to_excel(excel_path, sheet_name="Summary", index=False)

    # boxplot
    plt.figure()
    df[['aug_steps_init','aug_steps_opt']].boxplot()
    plt.title("Augmented Steps: Before vs After Optimization")
    plt.ylabel("Iterations to Converge")
    plot_path = out_dir/"boxplot.png"
    plt.tight_layout()
    plt.savefig(plot_path)
    plt.close()

    # 5) Open with openpyxl, insert matrices then the plot
    wb = load_workbook(excel_path)
    ws = wb["Summary"]

    # determine where the table ends (header row + df rows)
    table_end_row = 1 + len(df)  # header on row 1, data rows follow

    # we’ll start dumping matrices two rows below that
    current_row = table_end_row + 2

    for idx, A0 in enumerate(all_A0, start=1):
        # write a run header
        ws.cell(row=current_row, column=1).value = f"Run {idx} — initial A"
        # write the N×N block
        for i in range(N):
            for j in range(N):
                # round to 3 decimals
                ws.cell(row=current_row+1+i, column=1+j).value = float(f"{A0[i,j]:.3f}")

        current_row += N + 2


    img = XLImage(str(plot_path))
    ws.add_image(img, f"A{current_row}")

    wb.save(excel_path)
    print(f"Saved Excel (with tables, matrices, and boxplot) to {excel_path}")
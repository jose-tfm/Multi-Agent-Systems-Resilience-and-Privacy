import os
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import networkx as nx

from scipy.linalg import eigvals, eig
from scipy.optimize import minimize

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# ------------------------------------------------------------
# Row-stochastic normalization
def row_normalize(M: np.ndarray) -> np.ndarray:
    s = M.sum(axis=1, keepdims=True)
    s[s == 0] = 1e-12
    return M / s

# Build augmented P with fixed gadget weights
def build_Ap(N: int, A: np.ndarray) -> np.ndarray:
    size = 4 * N
    Ap = np.zeros((size, size))
    Ap[:N, :N] = A
    size = 4 * N
    Ap = np.zeros((size, size))
    Ap[:N, :N] = A  

    for i in range(N):
        inds = {
            0: i,          
            1: N + 3*i,   
            2: N + 3*i+1, 
            3: N + 3*i+2  
        }

        # ——————————————————————
        # a)
        
        Ap[inds[0], inds[1]] = 0.186
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[3]] = 0.108
        Ap[inds[3], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        
        # ——————————————————————
        # b)
        '''
        Ap[inds[0], inds[1]] = 0.01
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 0.01
        Ap[inds[0], inds[3]] = 0.134
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # c)
        '''
        Ap[inds[0], inds[1]] = 0.069
        Ap[inds[0], inds[2]] = 0.069
        Ap[inds[0], inds[3]] = 0.01
        Ap[inds[1], inds[3]] = 1
        Ap[inds[2], inds[3]] = 1
        Ap[inds[3], inds[0]] = 1
        '''
        # ——————————————————————
        # d)
        '''
        Ap[inds[0], inds[1]] = 0.01
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[3]] = 0.138
        Ap[inds[3], inds[0]] = 0.01
        Ap[inds[3], inds[2]] = 2.119
        Ap[inds[2], inds[0]] = 1
        '''
        # ——————————————————————
        # e)
        '''
        Ap[inds[0], inds[1]] = 0.134
        Ap[inds[1], inds[2]] = 1
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[2]] = 0.986
        Ap[inds[0], inds[3]] = 0.01
        Ap[inds[3], inds[0]] = 1.02
        '''
        # ——————————————————————
        # f)
        '''
        Ap[inds[0], inds[1]] = 0.069
        Ap[inds[1], inds[0]] = 0.01
        Ap[inds[1], inds[2]] = 2.109
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[3]] = 0.069
        Ap[inds[3], inds[0]] = 0.01
        Ap[inds[3], inds[2]] = 2.109
        '''
        # ——————————————————————
        # g)
        '''
        Ap[inds[0], inds[1]] = 0.01
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 0.01
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[3]] =  0.136
        Ap[inds[3], inds[0]] = 0.01
        Ap[inds[3], inds[2]] = 2.088
        '''
        # ——————————————————————
        # h)
        '''
        Ap[inds[0], inds[1]] = 0.01
        Ap[inds[1], inds[0]] = 1.069
        Ap[inds[1], inds[2]] = 0.932
        Ap[inds[2], inds[0]] = 1
        Ap[inds[0], inds[2]] = 0.01
        Ap[inds[0], inds[3]] = 0.132
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # i)
        '''
        Ap[inds[0], inds[1]] = 0.043
        Ap[inds[1], inds[2]] = 1
        Ap[inds[2], inds[1]] = 1.134
        Ap[inds[2], inds[0]] = 2.027
        Ap[inds[0], inds[2]] = 0.09
        Ap[inds[0], inds[3]] = 0.043
        Ap[inds[3], inds[2]] = 1
        '''
        # ——————————————————————
        # j)
        '''
        Ap[inds[0], inds[1]] = 0.01
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 0.135
        Ap[inds[2], inds[1]] = 1
        Ap[inds[2], inds[3]] = 1
        Ap[inds[0], inds[3]] = 0.01
        Ap[inds[3], inds[0]] = 1
        '''
        # ——————————————————————
        # k)
        '''
        Ap[inds[0], inds[1]] = 0.069
        Ap[inds[1], inds[0]] = 0.01
        Ap[inds[1], inds[2]] = 2.109
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[2]] = 2.109
        Ap[inds[0], inds[3]] = 0.069
        Ap[inds[3], inds[0]] = 0.01
        '''
        # ——————————————————————
        # l)
        '''
        Ap[inds[0], inds[1]] = 0.01
        Ap[inds[1], inds[0]] = 1
        Ap[inds[0], inds[2]] = 0.135
        Ap[inds[2], inds[0]] = 0.01
        Ap[inds[3], inds[0]] = 1
        Ap[inds[0], inds[3]] = 0.01
        Ap[inds[2], inds[1]] = 1.521
        Ap[inds[2], inds[3]] = 1.521
       '''
        # ——————————————————————
        # m)
        '''
        Ap[inds[0], inds[1]] = 0.069
        Ap[inds[1], inds[0]] = 0.01
        Ap[inds[0], inds[2]] = 0.01
        Ap[inds[2], inds[0]] = 1
        Ap[inds[3], inds[0]] = 0.01
        Ap[inds[0], inds[3]] = 0.069
        Ap[inds[1], inds[2]] = 2.139
        Ap[inds[3], inds[2]] = 2.139
        '''
    return row_normalize(Ap)

def build_A_from_p(p: np.ndarray, mask: np.ndarray) -> np.ndarray:
    """Reconstruct A (row-stochastic) from flattened free entries p."""
    N = mask.shape[0]
    A = np.zeros((N, N))
    idx = 0
    for i in range(N):
        for j in range(N):
            if mask[i, j]:
                A[i, j] = p[idx]
                idx += 1
    return row_normalize(A)

def simulate_and_metrics(A: np.ndarray,
                         x0: np.ndarray,
                         tol: float = 1e-4,
                         max_steps: int = 500
                        ) -> dict:
    """
    Simulate both x_{k+1}=A x_k and augmented x_{k+1}=P x_k,
    return lambda2, orig_steps and aug_steps.
    """
    N = A.shape[0]
    P = build_Ap(N, A)
    # λ₂
    ev = np.abs(eigvals(P))
    lambda2 = np.sort(ev)[-2]
    # build x_p0 that respects stationary left eigenvector
    w, V = eig(P.T)
    idx = np.argmin(np.abs(w - 1))
    v0 = np.real(V[:, idx])
    v0 /= v0.sum()
    target = x0.mean()
    x_p0 = np.zeros(4*N)
    for i in range(N):
        x_p0[N+3*i:N+3*i+3] = 4*x0[i]/3
    x_p0 *= target / (v0 @ x_p0)
    # simulate
    orig_steps = aug_steps = None
    xo = x0.copy()
    xa = x_p0.copy()
    for k in range(1, max_steps+1):
        xo = A @ xo
        xa = P @ xa
        if orig_steps is None and np.max(np.abs(xo - target)) < tol:
            orig_steps = k
        if aug_steps is None and np.max(np.abs(xa - target)) < tol:
            aug_steps = k
        if orig_steps is not None and aug_steps is not None:
            break
    return {
        'lambda2':     float(lambda2),
        'orig_steps':  orig_steps,
        'aug_steps':   aug_steps
    }

def make_obj(N: int, mask: np.ndarray):
    """Closure that returns λ₂(build_Ap(N, A(p))) for optimizer."""
    def obj(p):
        A = build_A_from_p(p, mask)
        return np.sort(np.abs(eigvals(build_Ap(N, A))))[-2]
    return obj

# -------------------------------------------------------------------
# 2) Main experiment loop + Excel export
# -------------------------------------------------------------------
if __name__ == "__main__":
    N        = 50     
    p_edge   = 0.3      
    num_runs = 70       
    rng      = np.random.RandomState(1234)
    x0_common = rng.rand(N)

    # storage
    summary = []
    all_A0  = []
    all_Aopt= []

    for run in range(1, num_runs+1):
        seed = 4 + run
        r_rng = np.random.RandomState(seed)

        while True:
            D = nx.gnp_random_graph(N,
                                    p_edge,
                                    seed=rng,
                                    directed=True)
            if nx.is_strongly_connected(D):
                G = D
                break
        # random positive weights
        for u, v in G.edges():
            G[u][v]['weight'] = r_rng.rand()
        # build A0
        raw_A = row_normalize(nx.to_numpy_array(G, nodelist=range(N), weight='weight'))
        mask  = raw_A > 0
        p0    = np.array([raw_A[i,j] for i in range(N) for j in range(N) if mask[i,j]])
        A0    = build_A_from_p(p0, mask)
        all_A0.append(A0)

        # simulate initial
        m0 = simulate_and_metrics(A0, x0_common, tol=1e-4, max_steps=500)

        # optimize
        res = minimize(
            make_obj(N, mask),
            p0,
            method='SLSQP',
            bounds=[(0.1, None)] * p0.size,
            options={'ftol':1e-6,'maxiter':300}
        )
        Aopt = build_A_from_p(res.x, mask)
        all_Aopt.append(Aopt)
        m1   = simulate_and_metrics(Aopt, x0_common, tol=1e-4, max_steps=500)

        # percentage speedup in augmented convergence
        if m0['aug_steps'] and m1['aug_steps']:
            pct = 100*(m0['aug_steps'] - m1['aug_steps'])/m0['aug_steps']
        else:
            pct = np.nan

        summary.append({
            'run':            run,
            'lambda2_init':   m0['lambda2'],
            'lambda2_opt':    m1['lambda2'],
            'aug_steps_init':  m0['aug_steps'],
            'aug_steps_opt':   m1['aug_steps'],
            'pct_speedup':     pct
        })

        print(
            f"Run {run}/{num_runs} — "
            f"λ_init={m0['lambda2']:.4f}, steps_init={m0['aug_steps']}; "
            f"λ_opt={m1['lambda2']:.4f}, steps_opt={m1['aug_steps']}; "
            f"speedup={pct:.1f}%"
        )

    # build DataFrame
    df = pd.DataFrame(summary)

    # prepare output folder
    out_dir = Path("ExcelDataRuns")
    out_dir.mkdir(exist_ok=True)

    # write summary to Excel
    excel_path = out_dir/f"Both_N=50_P=0.3.xlsx"
    df.to_excel(excel_path, sheet_name="Summary", index=False)

    # make boxplot of augmented steps
    plt.figure()
    df[['aug_steps_init','aug_steps_opt']].boxplot()
    plt.title("Augmented Convergence Steps: before vs after")
    plt.ylabel("Iterations to reach tol")
    box_png = out_dir/"boxplot.png"
    plt.tight_layout(); plt.savefig(box_png); plt.close()

    # insert the A₀ and Aopt matrices plus the plot
    wb = load_workbook(excel_path)
    ws = wb["Summary"]

    # after the table
    row0 = 1 + len(df) + 2
    for idx, (A0, Aopt) in enumerate(zip(all_A0, all_Aopt), start=1):
        # label + A₀
        ws.cell(row=row0, column=1).value = f"Run {idx} — A₀"
        for i in range(N):
            for j in range(N):
                ws.cell(row=row0+1+i, column=1+j).value = float(f"{A0[i,j]:.3f}")
        row0 += N+2
        # label + Aopt
        ws.cell(row=row0, column=1).value = f"Run {idx} — A_opt"
        for i in range(N):
            for j in range(N):
                ws.cell(row=row0+1+i, column=1+j).value = float(f"{Aopt[i,j]:.3f}")
        row0 += N+2

    # embed the boxplot
    img = XLImage(str(box_png))
    ws.add_image(img, f"A{row0}")

    wb.save(excel_path)
    print(f"\n✅ Full report saved to {excel_path}")